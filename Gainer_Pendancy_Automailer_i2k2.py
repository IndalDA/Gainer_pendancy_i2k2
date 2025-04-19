import os
import pyodbc
import time
import openpyxl
import pandas as pd
from io import BytesIO
import streamlit as st 
from datetime import datetime
from  openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.styles.alignment import Alignment
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles.colors import Color

#connection for db
conn = pyodbc.connect(
    r'DRIVER={ODBC Driver 17 for SQL Server};'
    r'SERVER=103.234.185.132,2499;'
    r'DATABASE=Z_SCOPE;'
    r'UID=Utkrishtsa;'
    r'PWD=AsknSDV*3h9*RFhkR9j73;')
cursor = conn.cursor()

# Function to convert DataFrame to Excel
def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        format1 = workbook.add_format({'num_format': '0.00'})
        worksheet.set_column('A:A', None, format1)
    return output.getvalue()


# Set page title and icon
st.set_page_config(page_title="Buying Potential Automailer", page_icon="📧")

#Backgound 
def set_bg_color():
    st.markdown(
        """
        <style>
        .stApp {
            background-color:#00BFBF ;  /* Light Blue */
            
        }
        </style>
        """,
        unsafe_allow_html=True
    )

# Call the function to set the background color
set_bg_color()


st.header('Buying Potential Automailer', divider="green")

loc_df = pd.read_sql_query('''select  a.Brand,a.Dealer,a.Location,concat(a.Brand,'_',a.dealer,'_',a.location) as select_location ,a.BrandID,a.DealerID,a.LocationID,
case when b.Consignee_Type='c' then 'Child'
when b.Consignee_Type='s' then 'Single'
when b.Consignee_Type='M' then 'Mother' end Consignee_Type
From locationinfo a
inner join Dealer_Setting_Master b on a.LocationID=b.locationid
where a.SharingStatus=1 and b.status=1 and ConsigneeType is not null
  and a.brandid in ('9','11','32','28')''',conn)

# Brand Dropdown
brnd = loc_df['Brand'].unique()
brand_list = ["Select Brand"] + brnd.tolist()
brand = st.selectbox(label="Brand", options=brand_list)


# Fetch dealer data based on selected brand
if brand:
    dealer = loc_df[loc_df['Brand']==brand]['Dealer'].unique()
    dealer_list = dealer.tolist()
    Dealer = st.multiselect("Select Dealer", options=dealer_list)
else:
    Dealer = []

col1,col2,col3,col4 = st.columns(4)
with col1:
    single = st.checkbox('Single')
with col2:    
    child = st.checkbox('Child')
with col3:    
    mother = st.checkbox('Mother')
with col4:
    All = st.checkbox('All')


# Show which options are selected
selected_types = []
if single:
    selected_types.append('Single')
if child:
    selected_types.append('Child')
if mother:
    selected_types.append('Mother')


# selected options
if selected_types:
    st.write(f"Selected options: {', '.join(selected_types)}")
else:
    st.write("No options selected")

# based on checkboxes
if selected_types:
    filtered_df =loc_df[loc_df['Consignee_Type'].isin(selected_types)]
else:
    filtered_df = loc_df  

#location data based on selected brand and dealers 
if brand != "Select Brand" and Dealer:
  
    filtered_df = filtered_df[filtered_df['Brand'] == brand]
    filtered_df = filtered_df[filtered_df['Dealer'].isin(Dealer)]  
    # selected brand, dealer, and consignee type
    location_list = filtered_df['select_location'].tolist()
    Location = st.multiselect("Location", options=location_list)
else:
    Location = []





#Parameter Based on Brand
if brand =="JCB":
    brandid = 32
    MinDis = 25
    OEM_Avg_Margin =18
    GST = float(0.18)
    Q1 = float(0.63)
    Q2 = float(0.83)
elif brand=="Mahindra":
    brandid = 9
    MinDis = 25
    OEM_Avg_Margin = 23
    GST = float(0.28)
    Q1 = float(0.63)
    Q2 = float(0.83)
elif brand=="TATA PCBU":
    brandid = 28
    MinDis = 25
    OEM_Avg_Margin = 23
    GST = float(0.28)
    Q1 = float(0.63)
    Q2 = float(0.83)
elif brand=='Hyundai':
    brandid = 11
    MinDis = 20
    OEM_Avg_Margin = 17
    GST = float(28)
    Q1 = float(0.63)
    Q2 = float(0.83)

if brand!="Select Brand":
        parameter = {'Brand':brand,'Minimum Discount':MinDis,'Oem Avg Margin':OEM_Avg_Margin,
                    'Gst':GST,'[Rate/MRP greater than]':Q1,'[Rate/MRP Less than]':Q2
                    }
        
#tabs craetion
#tab1, tab2, tab3, tab4 = st.tabs(['Parameter', 'Show_selection', 'Generate_Report', 'Mailsent'])
tab1, tab2, tab3, tab4,tab5 = st.tabs(['Parameter', 'Show_selection', 'Generate_Report','Summmary', 'Mailsent'])

#tab1, tab2, tab3, tab4,tab5 = st.tabs(['Parameter', 'Show_selection', 'Generate_Report','Summmary', 'Mailsent'])

# Tab 1: Parameter
with tab1:
    if brand != "Select Brand" and st.button(label="Click Here To Check Parameter", key=1):
        st.dataframe(pd.DataFrame(parameter, index=[0]))
    elif brand == "Select Brand" and st.button(label="Click Here To Check Parameter", key=1):
        st.warning('Select Brand')

# Tab 2: Show Selection
with tab2:
    if brand and Dealer and Location:
        try:
           
            Selected_df =loc_df[loc_df['select_location'].isin(Location)][['Brand','Dealer','Location','BrandID','DealerID','LocationID']]
            st.table(Selected_df)
        except Exception as e:
            st.error(f"Database error: {e}")
    else:
        st.warning('Error: Select Brand, Dealer, and Location!')

# Tab 3: Generate Report
with tab3:
    if brand and Dealer and Location:
        if st.button('Generate Report', key=2):
            #st.success("Report generated successfully!")
            for brndid,dlr,locid,brd,dlrn,locn in zip(Selected_df['BrandID'],Selected_df['DealerID'],Selected_df['LocationID'],
                                    Selected_df['Brand'],Selected_df['Dealer'],Selected_df['Location']):
                cursor.execute("exec USP_BuyingPotential_For_Automation1 ?,?,?,?,?,?", (brndid, dlr, locid, '1000', MinDis, '2'))
                conn.commit()
                gen_df = pd.read_sql_query('''select *from Uad_BuyingPotential_Automation where locationid = ?''',conn,params=(locid))

                st.success(f'Report Generate for : brand :{brd},Dealer :{dlrn},Location : {locn}')
        else:
            st.warning('Error: Select Brand, Dealer, and Location!')

with tab4:
    if st.button('Show Summary'):
        Selected_df =loc_df[loc_df['select_location'].isin(Location)][['Brand','Dealer','Location','BrandID','DealerID','LocationID']]
        Selected_df['Unque_Dealer'] = Selected_df['Brand'] + "_" + Selected_df['Dealer']+"_"+Selected_df['Location']
        #mail list 
        Mail_df = pd.read_csv(r'https://docs.google.com/spreadsheets/d/e/2PACX-1vRDqBXCxlSXSgOHUAUH6rPqtDQ-RWg9f0AOTFJH2-gAGOoJqubSFjGgRsJjmkECWyeWAP65Vx789z6B/pub?gid=1610467454&single=true&output=csv')
        Mail_df['unique_dealer'] = Mail_df['Brand'] + "_" + Mail_df['Dealer'] + "_" + Mail_df['Location']
        merge_df = Selected_df.merge(Mail_df, left_on='Unque_Dealer', right_on='unique_dealer', how='inner')
        
        report_gen =[]
        for brndid,dlr,locid,uniq in zip(merge_df['BrandID'],merge_df['DealerID'],merge_df['LocationID'],merge_df['Unque_Dealer']):
            print((brndid, dlr, locid, Q1, Q2, GST, OEM_Avg_Margin))
            cursor.execute("uad_buyingPotential_automation_SP_Msg_pra_vs_test ?,?,?,?,?,?,?",(brndid, dlr, locid, Q1, Q2, GST, OEM_Avg_Margin))
            cursor.commit() 
            report_gen.append(str(locid))

        loc_ids_str = ','.join(report_gen)

        if loc_ids_str:
            summary_query = f'''
                select b.Brand,a.* FROM uad_buyingPotential_automation_SP_whatsapp a
                inner join locationinfo b on a.locationid=b.locationid
                WHERE CAST(ReportGeneratedOn AS DATE) = CAST(GETDATE() AS DATE)
                    AND a.LocationID IN ({loc_ids_str})
            '''
        
            bo_data = pd.read_sql_query(summary_query, conn)
            summary = bo_data.groupby(['DEALER', 'Location']).agg({'PartNumber':'count','Disc_value':'sum','Add_Profit':'sum'}).reset_index()
            summary.columns=['Dealer', 'Location', 'Part_Count', 'Buying Opp Value (in Rs)','Est Addl Profit (in Rs)']
            summary['Buying Opp Value (in Rs)']=summary['Buying Opp Value (in Rs)'].astype(int)
            summary['Est Addl Profit (in Rs)']=summary['Est Addl Profit (in Rs)'].astype(int)    
            
            #st.table(summary)
            if len(summary)>0:
                st.data_editor(summary,num_rows='dynamic')
            else:
                st.warning('Buying Potential not Generated:',icon="ℹ️")   
            # export data     
            Mail_df = pd.read_csv(r'https://docs.google.com/spreadsheets/d/e/2PACX-1vRDqBXCxlSXSgOHUAUH6rPqtDQ-RWg9f0AOTFJH2-gAGOoJqubSFjGgRsJjmkECWyeWAP65Vx789z6B/pub?gid=1610467454&single=true&output=csv')
            Mail_df['unique_dealer'] = Mail_df['Brand'] + "_" + Mail_df['Dealer'] + "_" + Mail_df['Location']
            bo_data['Unque_Dealer'] = bo_data['Brand']+"_"+bo_data['DEALER']+"_"+bo_data['Location']
            own_df = bo_data.merge(Mail_df, left_on='Unque_Dealer', right_on='unique_dealer', how='left')
            df_xlsx = to_excel(own_df)                
            st.download_button(
                label="📥Download Bo Excel File",
                data=df_xlsx,
                file_name=f"{brand}_Buying_Potential.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            print("No valid Location IDs to query.")
    # if st.button('Show Summary'):
    #         Selected_df =loc_df[loc_df['select_location'].isin(Location)][['Brand','Dealer','Location','BrandID','DealerID','LocationID']]
    #         Selected_df['Unque_Dealer'] = Selected_df['Brand'] + "_" + Selected_df['Dealer']+"_"+Selected_df['Location']
    #         #mail list 
    #         Mail_df = pd.read_csv(r'https://docs.google.com/spreadsheets/d/e/2PACX-1vRDqBXCxlSXSgOHUAUH6rPqtDQ-RWg9f0AOTFJH2-gAGOoJqubSFjGgRsJjmkECWyeWAP65Vx789z6B/pub?gid=1610467454&single=true&output=csv')
    #         Mail_df['unique_dealer'] = Mail_df['Brand'] + "_" + Mail_df['Dealer'] + "_" + Mail_df['Location']
    #         merge_df = Selected_df.merge(Mail_df, left_on='Unque_Dealer', right_on='unique_dealer', how='inner')
            
    #         report_gen =[]
    #         for brndid,dlr,locid,uniq in zip(merge_df['BrandID'],merge_df['DealerID'],merge_df['LocationID'],merge_df['Unque_Dealer']):
    #             print((brndid, dlr, locid, Q1, Q2, GST, OEM_Avg_Margin))
    #             cursor.execute("delete from uad_buyingPotential_automation_SP_whatsapp where  cast(ReportGeneratedOn as date)= cast(getdate() as date) and locationid=?",(locid))        
    #             cursor.commit()
    #             cursor.execute("uad_buyingPotential_automation_SP_Msg_pra_vs_test ?,?,?,?,?,?,?",(brndid, dlr, locid, Q1, Q2, GST, OEM_Avg_Margin))
    #             cursor.commit() 
    #             report_gen.append(str(locid))

    #         loc_ids_str = ','.join(report_gen)

    #         if loc_ids_str:
    #             summary_query = f'''
    #                 SELECT Dealer, Location, COUNT(PartNumber) AS Part_Count,
    #                     round(SUM(Mrp_value),0) AS [Buying Opp Value (in Rs)],
    #                     round(SUM(Add_Profit),0) AS [Est Addl Profit (in Rs)]
    #                 FROM uad_buyingPotential_automation_SP_whatsapp
    #                 WHERE CAST(ReportGeneratedOn AS DATE) = CAST(GETDATE() AS DATE)
    #                     AND LocationID IN ({loc_ids_str})
    #                 GROUP BY DEALER, Location
    #             '''
            
    #             summary = pd.read_sql_query(summary_query, conn)
    #             summary['Buying Opp Value (in Rs)']=summary['Buying Opp Value (in Rs)'].astype(int)
    #             summary['Est Addl Profit (in Rs)']=summary['Est Addl Profit (in Rs)'].astype(int)
    #             #st.table(summary)
    #             if len(summary)>0:
    #                 st.data_editor(summary,num_rows='dynamic')
    #             else:
    #                 st.warning('Buying Potential not Generated:',icon="ℹ️")     
    #         else:
    #             print("No valid Location IDs to query.")

# Tab 4: Mail Sent
with tab5:
    import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime
import smtplib as s
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import time
import os
import pathlib

# Your existing pre-loop logic
if brand and Dealer and Location:
    if st.button('Click Here to Sent Report', key=3):
        Selected_df = loc_df[loc_df['select_location'].isin(Location)][['Brand','Dealer','Location','BrandID','DealerID','LocationID']]
        Selected_df['Unque_Dealer'] = Selected_df['Brand'] + "_" + Selected_df['Dealer']+"_"+Selected_df['Location']

        Mail_df = pd.read_csv('https://docs.google.com/spreadsheets/d/e/2PACX-1vRDqBXCxlSXSgOHUAUH6rPqtDQ-RWg9f0AOTFJH2-gAGOoJqubSFjGgRsJjmkECWyeWAP65Vx789z6B/pub?gid=1610467454&single=true&output=csv')
        Mail_df['unique_dealer'] = Mail_df['Brand'] + "_" + Mail_df['Dealer'] + "_" + Mail_df['Location']

        merge_df = Selected_df.merge(Mail_df, left_on='Unque_Dealer', right_on='unique_dealer', how='inner')

        # 🔍 Search functionality
        search_query = st.text_input("🔍 Search by Brand or Dealer", "")
        filtered_df = merge_df[
            merge_df['Brand'].str.contains(search_query, case=False, na=False) |
            merge_df['Dealer'].str.contains(search_query, case=False, na=False)
        ]

        # Prepare summary data for editing
        summary_data = []
        for _, row in filtered_df.iterrows():
            locid = row['LocationID']
            q_df = pd.read_sql('''SELECT [Buying Opportunity Qty], Disc_value, Add_Profit, PartNumber
                                  FROM uad_buyingPotential_automation_SP_whatsapp 
                                  WHERE [Buying Opportunity Qty] > 0 
                                  AND Locationid = ? 
                                  AND CAST(ReportGeneratedOn AS date) = CAST(GETDATE() AS date)''',
                                  conn, params=(locid,))
            
            summary_data.append({
                'Brand': row['Brand'],
                'Dealer': row['Dealer'],
                'Location': row['Location_x'],
                'BrandID': row['BrandID'],
                'DealerID': row['DealerID'],
                'LocationID': row['LocationID'],
                'Unque_Dealer': row['Unque_Dealer'],
                'Part Count': q_df['PartNumber'].nunique(),
                'Buying Opportunity Qty': q_df['Buying Opportunity Qty'].sum(),
                'Disc_value': int(q_df['Disc_value'].sum()),
                'Add_Profit': int(q_df['Add_Profit'].sum()),
                'To': row['To'],
                'CC': row['CC'],
                'Send Email': True
            })

        editable_df = pd.DataFrame(summary_data)

        # Session state to store editable table
        if 'edited_data' not in st.session_state:
            st.session_state.edited_data = editable_df.copy()

        # Editable table
        st.markdown("### ✏️ Edit Email Info or Deselect to Skip")
        st.session_state.edited_data = st.data_editor(
            st.session_state.edited_data,
            num_rows="dynamic",
            use_container_width=True,
            key="editable_email_table"
        )

        # Confirmation modal simulation
        if st.button("✅ Confirm & Send Emails"):
            st.session_state.show_confirm = True

        if st.session_state.get("show_confirm"):
            with st.expander("📨 Final Confirmation — Click 'Send Now' to dispatch emails"):
                if st.button("🚀 Send Now"):
                    for _, row in st.session_state.edited_data.iterrows():
                        if not row['Send Email']:
                            continue

                        brndid = row['BrandID']
                        dlr = row['DealerID']
                        locid = row['LocationID']
                        uniq = row['Unque_Dealer']
                        to_email = row['To']
                        cc_emails = row['CC']
                        Location_name = row['Location']

                        cursor.execute("delete from uad_buyingPotential_automation_SP_whatsapp where Locationid =? and cast(ReportGeneratedOn as date)= cast(getdate() as date)",(locid,))        
                        cursor.commit()

                        cursor.execute("uad_buyingPotential_automation_SP_Msg_pra_vs_test ?,?,?,?,?,?,?",(brndid, dlr, locid, Q1, Q2, GST, OEM_Avg_Margin))
                        cursor.commit()

                        df2 = pd.read_sql('''select distinct DEALER, Location, PartNumber, PartDescription, Mrp, Rate,     
                                    [3M Avg Sale], [Gainer Free Stock], [Buying Opportunity Qty], Discount,     
                                    Disc_value, Mrp_value, Add_Profit    
                                    from uad_buyingPotential_automation_SP_whatsapp where [Buying Opportunity Qty] >0 and Locationid =? and 
                                    cast(ReportGeneratedOn as date)= cast(getdate() as date)''', conn,params=(locid,))

                        if len(df2) == 0:
                            st.warning(f'No data available for: {uniq}')
                            continue

                        Part_count = df2['PartNumber'].count()
                        Pur_Value = int(round(df2['Disc_value'].sum(), 0))
                        Add_Profit = int(round(df2['Add_Profit'].sum(), 0))

                        # Save Excel
                        Document_folder = str(pathlib.Path.home() / "Documents")
                        output_excel = Document_folder + "\\Buying Potential For " + uniq + " " + datetime.strftime(datetime.now(), '%Y-%m-%d') + ".xlsx"
                        df2.to_excel(output_excel, index=False)

                        wb = load_workbook(output_excel)
                        ws = wb.active
                        ws.alignment = Alignment(horizontal='center', vertical='center')
                        for cell in ws[1]:
                            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill('solid', start_color="38ffe9")
                        max_row = ws.max_row
                        max_col = ws.max_column
                        for row_idx in range(1, max_row + 1):
                            for col_idx in range(1, max_col + 1):
                                cell = ws.cell(row=row_idx, column=col_idx)
                                cell.border = Border(left=Side(border_style='thin', color='000000'),
                                                     right=Side(border_style='thin', color='000000'),
                                                     top=Side(border_style='thin', color='000000'),
                                                     bottom=Side(border_style='thin', color='000000'))
                        wb.save(output_excel)

                        # Prepare email
                        body = f'''
                        <p style="font-family: 'Calibri', Times, serif;">Dear Sir,</p>
                        <p>Please find attached list of Spare Parts available on Sparecare Gainer Portal at <b>HIGH DISCOUNTS ~ 25-50%</b>.</p>
                        <p><b>
                        Location Name           : {Location_name} <br>
                        Parts Count             : {Part_count} Nos <br>
                        Parts Discounted Value  : Rs {Pur_Value}/- <br>
                        Addl. Profit*           : Rs {Add_Profit}/-
                        </b></p>
                        <p style="color: blue;">* above OE Margin</p>
                        <p>Kindly check the list & place order on Gainer Portal.<br>
                        For any support required, please revert.</p>
                        <p>Thanks & Regards <br>Team Gainer</p>
                        '''

                        cc_emails = cc_emails.replace(' ', '')
                        cc_email_list = [email for email in cc_emails.split(';') if email]
                        all_recipients = [to_email] + cc_email_list

                        msg = MIMEMultipart()
                        msg['From'] = 'gainer.alerts@sparecare.in'
                        msg['To'] = to_email
                        msg['Cc'] = ', '.join(cc_email_list)
                        msg['Subject'] = f'Addl Profit Opportunity (Buying) :- {uniq}'
                        msg.attach(MIMEText(body, 'html'))

                        with open(output_excel, "rb") as attachment:
                            part = MIMEBase('application', 'octet-stream')
                            part.set_payload(attachment.read())
                            encoders.encode_base64(part)
                            part.add_header('Content-Disposition', f"attachment; filename= {output_excel.split('/')[-1]}")
                            msg.attach(part)

                        ob = s.SMTP('smtp.gmail.com', 587)
                        ob.ehlo()
                        ob.starttls()
                        ob.login('gainer.alerts@sparecare.in', 'fmyclggqzrmkykol')
                        ob.sendmail('gainer.alerts@sparecare.in', all_recipients, msg.as_string())
                        ob.quit()
                        st.success("✅ Email sent for: " + uniq)
    else:
        st.warning('Error: Select Brand, Dealer, and Location!')
   


cursor.close()
conn.close()
